import pandas as pd
from openpyxl import load_workbook

def read_excel(path: str):
    """"
    Lê planilhas Excel e extrai metadados e amostras de dados.
    """

    sheet_df = pd.read_excel(path, sheet_name="Cert Mileage Forecast - All", skiprows=4, index_col=0)
    sheet_df.columns = sheet_df.columns.str.strip()
    ignore = ['DYNO CONSTRAINT (2WD/AWD)', 'GAS TYPE (ONLY FOR CERT)', 'CYCLE', 'NOTES']
    sheet_df.drop(columns=ignore, inplace=True)
    sheet_df.reset_index(inplace=True)
    
    return {
        "Cert Mileage Forecast - All": {
            "columns": sheet_df.columns.tolist(),
            "types": sheet_df.dtypes.astype(str).to_dict(),
            "samples": sheet_df.head(10).to_dict(orient="records"),
            "num_rows": len(sheet_df),
            "data": sheet_df #.to_dict(orient="records")
        }
    }


def apply_actions(path: str, actions: list):
    """
    Aplica ações em uma planilha Excel (.xlsm) preservando macros VBA.

    Parâmetros:
        path (str): Caminho do arquivo Excel (.xlsm)
        actions (list): Lista de ações JSON conforme especificação do agente

    Ações suportadas:
        - update_cell
        - add_row
        - delete_row
        - update_column
        - add_column
        - delete_column
    """

    wb = load_workbook(path, keep_vba=True)
    sheet_names = wb.sheetnames

    for act in actions:
        action = act.get("action")
        sheet = act.get("sheet", "Sheet1")

        if sheet not in sheet_names:
            raise ValueError(f"Aba '{sheet}' não encontrada no arquivo.")

        ws = wb[sheet]

        # ========== UPDATE CELL ==========
        if action == "update_cell":
            row = act.get("row")
            col = act.get("col")
            value = act.get("value")

            if isinstance(col, str):
                headers = [cell.value for cell in ws[1]]
                if col in headers:
                    col = headers.index(col) + 1
                else:
                    raise ValueError(f"Coluna '{col}' não encontrada na planilha {sheet}.")

            ws.cell(row=row, column=col, value=value)

        # ========== ADD ROW ==========
        elif action == "add_row":
            values = act.get("values", {})
            headers = [cell.value for cell in ws[1]]
            new_row = [values.get(h, None) for h in headers]
            ws.append(new_row)

        # ========== DELETE ROW ==========
        elif action == "delete_row":
            row = act.get("row")
            ws.delete_rows(row)

        # ========== UPDATE COLUMN (with condition) ==========
        elif action == "update_column":
            column = act.get("column")
            new_value = act.get("new_value")
            condition = act.get("condition", "")
            headers = [cell.value for cell in ws[1]]

            if column not in headers:
                raise ValueError(f"Coluna '{column}' não encontrada.")

            col_idx = headers.index(column) + 1

            # Avaliar a condição linha a linha
            for row in range(2, ws.max_row + 1):
                row_values = {headers[i]: ws.cell(row=row, column=i + 1).value for i in range(len(headers))}
                try:
                    if eval(condition, {}, row_values):
                        ws.cell(row=row, column=col_idx, value=new_value)
                except Exception:
                    continue

        # ========== ADD COLUMN ==========
        elif action == "add_column":
            column_name = act.get("column_name")
            default_value = act.get("default_value", None)
            headers = [cell.value for cell in ws[1]]

            if column_name in headers:
                continue

            ws.cell(row=1, column=len(headers) + 1, value=column_name)
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=len(headers) + 1, value=default_value)

        # ========== DELETE COLUMN ==========
        elif action == "delete_column":
            column_name = act.get("column_name")
            headers = [cell.value for cell in ws[1]]

            if column_name not in headers:
                continue

            idx = headers.index(column_name) + 1
            ws.delete_cols(idx)

        else:
            print(f"[WARN] Ação desconhecida ignorada: {action}")

    wb.save(path)
    wb.close()
    print(f"[INFO] Ações aplicadas com sucesso em: {path}")